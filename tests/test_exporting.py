from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

from hospital_ocr.consolidation import consolidate_records
from hospital_ocr.exporting import export_results
from hospital_ocr.models import PatientRecord


def test_excel_contains_required_sheets_and_template(tmp_path: Path) -> None:
    patient = PatientRecord(
        full_name="María Pérez",
        first_name="María",
        last_name="Pérez",
        name_split_confidence=1.0,
        detected_name_order="Nombre-Apellido",
        center="Hospital de Prueba",
        age=8,
        age_unit="años",
        sex="F",
        origin="Petare",
        specialty="Pediatría",
        area="UCI",
        source_image="lista.jpg",
        confidence=0.95,
        needs_review=False,
        raw_line="María Pérez 8a F Petare",
        clinical_notes="Plan: Politrauma",
    )
    output = tmp_path / "pacientes.xlsx"

    export_results(
        consolidate_records([patient]),
        output,
        specialty_values=["Pediatría"],
    )

    workbook = load_workbook(output)
    assert workbook.sheetnames == [
        "Plantilla",
        "Pacientes",
        "Diccionario",
    ]
    assert workbook["Plantilla"].sheet_state == "visible"
    assert workbook["Pacientes"].sheet_state == "visible"

    template = pd.read_excel(output, sheet_name="Plantilla")
    assert list(template.columns) == [
        "nombre",
        "apellido",
        "cedula",
        "centro",
        "edad_sector",
    ]
    template_sheet = workbook["Plantilla"]
    assert template_sheet.max_row == 1001
    assert template_sheet["A2"].value.startswith("=IF(Pacientes!")
    assert 'IF(Pacientes!$G2="","",Pacientes!$G2)' in template_sheet["C2"].value
    assert "Pacientes!" in template_sheet["E2"].value
    assert template_sheet["A1"].comment is not None
    assert template_sheet.protection.sheet is True
    assert not (tmp_path / "plantilla_pacientes.csv").exists()

    patients = pd.read_excel(output, sheet_name="Pacientes")
    assert "cedula" in patients.columns
    assert "nombre_completo" in patients.columns
    assert "nombre_completo_ocr" not in patients.columns
    assert "confianza_separacion_nombre" in patients.columns
    assert "orden_nombre_detectado" in patients.columns
    assert "confianza_ocr" in patients.columns
    assert "confianza_nombre" in patients.columns
    assert "confianza_cedula" in patients.columns
    assert "confianza_edad" in patients.columns
    assert "confianza_procedencia" in patients.columns
    assert "confianza_especialidad" in patients.columns
    assert "evidencia_extraccion" in patients.columns
    assert "linea_ocr_original" in patients.columns
    assert "imagenes_origen" in patients.columns
    assert "imagen_origen_principal" not in patients.columns
    assert "estado_revision" in patients.columns
    assert "requiere_revision" not in patients.columns
    assert patients.loc[0, "estado_revision"] == "No requerido"
    assert patients.loc[0, "observaciones"] == "Plan: Politrauma"
    assert "estado_duplicado" in patients.columns
    assert "detalle_duplicado" in patients.columns
    patient_sheet = workbook["Pacientes"]
    assert "TablaPacientes" in patient_sheet.tables
    assert len(patient_sheet.data_validations.dataValidation) == 4
    assert len(patient_sheet.conditional_formatting) == 1
    assert workbook.calculation.calcMode == "auto"
    assert workbook.calculation.fullCalcOnLoad is True

    dictionary = pd.read_excel(output, sheet_name="Diccionario")
    assert list(dictionary.columns) == [
        "hoja",
        "columna",
        "descripcion",
        "tipo",
        "valores_permitidos",
        "ejemplo",
    ]
    assert "nombre_completo" in dictionary["columna"].values
    assert set(
        dictionary.loc[dictionary["hoja"] == "Pacientes", "columna"]
    ) == set(patients.columns)
    assert set(
        dictionary.loc[dictionary["hoja"] == "Plantilla", "columna"]
    ) == set(template.columns)
