from __future__ import annotations

from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook
from openpyxl.comments import Comment
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter

from hospital_ocr.models import ConsolidationResult, PatientRecord


FULL_COLUMNS = [
    "id_paciente",
    "nombre_completo",
    "nombre",
    "apellido",
    "confianza_separacion_nombre",
    "orden_nombre_detectado",
    "cedula",
    "centro",
    "edad",
    "unidad_edad",
    "sexo",
    "procedencia",
    "especialidad",
    "area",
    "imagenes_origen",
    "confianza_ocr",
    "confianza_nombre",
    "confianza_cedula",
    "confianza_edad",
    "confianza_procedencia",
    "confianza_especialidad",
    "evidencia_extraccion",
    "estado_revision",
    "observaciones",
    "apariciones",
    "estado_duplicado",
    "detalle_duplicado",
    "linea_ocr_original",
]
TEMPLATE_COLUMNS = ["nombre", "apellido", "cedula", "centro", "edad_sector"]
TEMPLATE_CAPACITY = 1000
DICTIONARY_COLUMNS = [
    "hoja",
    "columna",
    "descripcion",
    "tipo",
    "valores_permitidos",
    "ejemplo",
]


DICTIONARY_ROWS = [
    ("Plantilla", "nombre", "Nombre o nombres confirmados.", "Texto", "", "María"),
    ("Plantilla", "apellido", "Apellido o apellidos confirmados.", "Texto", "", "Pérez"),
    ("Plantilla", "cedula", "Documento de identidad, si está disponible.", "Texto", "", "V-12345678"),
    ("Plantilla", "centro", "Nombre oficial del centro hospitalario.", "Texto", "", "Hospital de Prueba"),
    ("Plantilla", "edad_sector", "Edad y procedencia combinadas.", "Texto", "", "8 años · Petare"),
    ("Pacientes", "id_paciente", "Identificador interno del registro consolidado.", "Texto", "PAC-####", "PAC-0001"),
    ("Pacientes", "nombre_completo", "Nombre completo extraído de la imagen, sin separar cuando no hay certeza.", "Texto", "", "María Pérez"),
    ("Pacientes", "nombre", "Nombre o nombres; queda vacío si la separación no es confiable.", "Texto", "", "María"),
    ("Pacientes", "apellido", "Apellido o apellidos; queda vacío si la separación no es confiable.", "Texto", "", "Pérez"),
    ("Pacientes", "confianza_separacion_nombre", "Confianza del clasificador al separar nombres y apellidos.", "Decimal", "0 a 1", "0.95"),
    ("Pacientes", "orden_nombre_detectado", "Orden detectado en el texto original.", "Categoría", "Nombre-Apellido; Apellido-Nombre; Indeterminado", "Nombre-Apellido"),
    ("Pacientes", "cedula", "Documento de identidad; permanece vacío si no aparece.", "Texto", "", "V-12345678"),
    ("Pacientes", "centro", "Centro determinado por la carpeta de la imagen.", "Texto", "", "Hospital de Prueba"),
    ("Pacientes", "edad", "Valor numérico de la edad.", "Entero", "0 a 115", "8"),
    ("Pacientes", "unidad_edad", "Unidad asociada a la edad.", "Categoría", "años; meses; días", "años"),
    ("Pacientes", "sexo", "Sexo escrito explícitamente en la lista.", "Categoría", "M; F; vacío", "F"),
    ("Pacientes", "procedencia", "Sector o lugar de procedencia reconocido.", "Texto", "", "Petare"),
    ("Pacientes", "especialidad", "Especialidad hospitalaria normalizada.", "Texto", "", "Pediatría"),
    ("Pacientes", "area", "Área, piso o unidad dentro de la especialidad.", "Texto", "", "UCI"),
    ("Pacientes", "imagenes_origen", "Archivos donde apareció el paciente.", "Texto", "Nombres separados por punto y coma", "lista_01.jpg"),
    ("Pacientes", "confianza_ocr", "Confianza general del texto reconocido.", "Decimal", "0 a 1", "0.92"),
    ("Pacientes", "confianza_nombre", "Confianza de que el texto fue clasificado como nombre.", "Decimal", "0 a 1", "0.94"),
    ("Pacientes", "confianza_cedula", "Confianza de la cédula según formato y columna.", "Decimal", "0 a 1", "0.96"),
    ("Pacientes", "confianza_edad", "Confianza de la edad según formato, unidad y columna.", "Decimal", "0 a 1", "0.91"),
    ("Pacientes", "confianza_procedencia", "Confianza de la procedencia según columna y catálogo.", "Decimal", "0 a 1", "0.90"),
    ("Pacientes", "confianza_especialidad", "Confianza de la especialidad según membrete, columna y catálogo.", "Decimal", "0 a 1", "0.88"),
    ("Pacientes", "evidencia_extraccion", "Señales utilizadas para clasificar cada campo.", "Texto", "", "edad: formato y encabezado"),
    ("Pacientes", "estado_revision", "Indica si el registro necesita comprobación humana.", "Categoría", "Pendiente; No requerido; Revisado", "Pendiente"),
    ("Pacientes", "observaciones", "Plan indicado en la tabla, motivos de revisión y conflictos detectados.", "Texto", "", "Plan: Politrauma"),
    ("Pacientes", "apariciones", "Cantidad de registros fusionados.", "Entero", "1 o mayor", "2"),
    ("Pacientes", "estado_duplicado", "Resultado del análisis de duplicados.", "Categoría", "Único; Posible duplicado; Duplicado consolidado", "Único"),
    ("Pacientes", "detalle_duplicado", "Explicación y registros relacionados con el duplicado.", "Texto", "", "Coincide con PAC-0002 | nombre 93%"),
    ("Pacientes", "linea_ocr_original", "Línea completa producida por el OCR antes del análisis.", "Texto", "", "María Pérez 8a F Petare"),
]


def _safe_excel_text(value: str) -> str:
    if value.startswith(("=", "+", "-", "@")):
        return f"'{value}"
    return value


def _full_row(record: PatientRecord) -> dict[str, Any]:
    return {
        "id_paciente": record.patient_id,
        "nombre_completo": _safe_excel_text(record.full_name),
        "nombre": _safe_excel_text(record.first_name),
        "apellido": _safe_excel_text(record.last_name),
        "confianza_separacion_nombre": record.name_split_confidence,
        "orden_nombre_detectado": record.detected_name_order,
        "cedula": _safe_excel_text(record.document_id),
        "centro": record.center,
        "edad": record.age,
        "unidad_edad": record.age_unit,
        "sexo": record.sex,
        "procedencia": _safe_excel_text(record.origin),
        "especialidad": record.specialty,
        "area": record.area,
        "imagenes_origen": record.source_images_text,
        "confianza_ocr": record.confidence,
        "confianza_nombre": record.name_confidence,
        "confianza_cedula": record.document_confidence,
        "confianza_edad": record.age_confidence,
        "confianza_procedencia": record.origin_confidence,
        "confianza_especialidad": record.specialty_confidence,
        "evidencia_extraccion": _safe_excel_text(record.field_evidence_text),
        "estado_revision": record.review_status
        or ("Pendiente" if record.needs_review else "No requerido"),
        "observaciones": record.observations_text,
        "apariciones": record.occurrences,
        "estado_duplicado": record.duplicate_status,
        "detalle_duplicado": record.duplicate_detail,
        "linea_ocr_original": _safe_excel_text(record.raw_line),
    }


def patient_records_dataframe(records: list[PatientRecord]) -> pd.DataFrame:
    return pd.DataFrame(
        [_full_row(record) for record in records],
        columns=FULL_COLUMNS,
    )


def _header_columns(sheet: Any) -> dict[str, str]:
    return {
        str(cell.value): get_column_letter(cell.column)
        for cell in sheet[1]
        if cell.value
    }


def _link_template_sheet(workbook: Any) -> None:
    template = workbook["Plantilla"]
    patients = workbook["Pacientes"]
    patient_columns = _header_columns(patients)
    id_column = patient_columns["id_paciente"]
    mappings = {
        "A": patient_columns["nombre"],
        "B": patient_columns["apellido"],
        "C": patient_columns["cedula"],
        "D": patient_columns["centro"],
    }
    age_column = patient_columns["edad"]
    age_unit_column = patient_columns["unidad_edad"]
    origin_column = patient_columns["procedencia"]

    for row in range(2, TEMPLATE_CAPACITY + 2):
        for target_column, source_column in mappings.items():
            template[f"{target_column}{row}"] = (
                f'=IF(Pacientes!${id_column}{row}="","",'
                f'Pacientes!${source_column}{row})'
            )
        template[f"E{row}"] = (
            f'=IF(Pacientes!${id_column}{row}="","",'
            f'IF(Pacientes!${age_column}{row}="","",'
            f'Pacientes!${age_column}{row}&" "&Pacientes!${age_unit_column}{row})'
            f'&IF(AND(Pacientes!${age_column}{row}<>"",'
            f'Pacientes!${origin_column}{row}<>"")," · ","")'
            f'&Pacientes!${origin_column}{row})'
        )

    template["A1"].comment = Comment(
        "Hoja automática. Edite los datos únicamente en la hoja Pacientes.",
        "Hospital OCR",
    )
    template.protection.sheet = True
    template.sheet_properties.tabColor = "70AD47"
    template.column_dimensions["A"].width = 24
    template.column_dimensions["B"].width = 28
    template.column_dimensions["C"].width = 18
    template.column_dimensions["D"].width = 32
    template.column_dimensions["E"].width = 36


def _add_patient_table(sheet: Any) -> None:
    if sheet.max_row < 2:
        return
    table = Table(displayName="TablaPacientes", ref=sheet.dimensions)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    sheet.add_table(table)


def _add_list_validation(
    sheet: Any,
    column: str,
    values: list[str],
) -> None:
    if not values:
        return
    formula = '"' + ",".join(value.replace('"', '""') for value in values) + '"'
    validation = DataValidation(
        type="list",
        formula1=formula,
        allow_blank=True,
    )
    validation.error = "Seleccione un valor de la lista."
    validation.errorTitle = "Valor no válido"
    validation.showErrorMessage = True
    sheet.add_data_validation(validation)
    validation.add(f"{column}2:{column}{TEMPLATE_CAPACITY + 1}")


def _configure_patient_sheet(sheet: Any, specialty_values: list[str]) -> None:
    columns = _header_columns(sheet)
    _add_patient_table(sheet)
    _add_list_validation(sheet, columns["sexo"], ["M", "F"])
    _add_list_validation(sheet, columns["unidad_edad"], ["años", "meses", "días"])
    _add_list_validation(
        sheet,
        columns["estado_revision"],
        ["Pendiente", "No requerido", "Revisado"],
    )
    _add_list_validation(sheet, columns["especialidad"], specialty_values)

    last_column = get_column_letter(sheet.max_column)
    target_range = f"A2:{last_column}{TEMPLATE_CAPACITY + 1}"
    duplicate_column = columns["estado_duplicado"]
    review_column = columns["estado_revision"]
    sheet.conditional_formatting.add(
        target_range,
        FormulaRule(
            formula=[f'${duplicate_column}2="Posible duplicado"'],
            fill=PatternFill("solid", fgColor="FCE4D6"),
            stopIfTrue=True,
        ),
    )
    sheet.conditional_formatting.add(
        target_range,
        FormulaRule(
            formula=[f'${duplicate_column}2="Duplicado consolidado"'],
            fill=PatternFill("solid", fgColor="DDEBF7"),
            stopIfTrue=True,
        ),
    )
    sheet.conditional_formatting.add(
        target_range,
        FormulaRule(
            formula=[f'${review_column}2="Pendiente"'],
            fill=PatternFill("solid", fgColor="FFF2CC"),
        ),
    )
    sheet.sheet_properties.tabColor = "5B9BD5"


def _format_workbook(path: Path, specialty_values: list[str]) -> None:
    workbook = load_workbook(path)
    header_fill = PatternFill("solid", fgColor="1F4E78")
    _link_template_sheet(workbook)
    _configure_patient_sheet(workbook["Pacientes"], specialty_values)
    for sheet in workbook.worksheets:
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        for cell in sheet[1]:
            cell.fill = header_fill
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(horizontal="center")
        for column_cells in sheet.columns:
            values = [str(cell.value or "") for cell in column_cells]
            width = min(max(max(map(len, values), default=0) + 2, 10), 50)
            sheet.column_dimensions[column_cells[0].column_letter].width = width
        sheet.sheet_view.showGridLines = False
    template = workbook["Plantilla"]
    template.column_dimensions["A"].width = 24
    template.column_dimensions["B"].width = 28
    template.column_dimensions["C"].width = 18
    template.column_dimensions["D"].width = 32
    template.column_dimensions["E"].width = 36
    workbook.calculation.calcMode = "auto"
    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True
    workbook.save(path)


def export_results(
    result: ConsolidationResult,
    output_path: Path,
    specialty_values: list[str] | None = None,
) -> Path:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    records = result.patients
    patients_df = patient_records_dataframe(records)
    template_df = pd.DataFrame(columns=TEMPLATE_COLUMNS)
    dictionary_df = pd.DataFrame(DICTIONARY_ROWS, columns=DICTIONARY_COLUMNS)

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        template_df.to_excel(writer, sheet_name="Plantilla", index=False)
        patients_df.to_excel(writer, sheet_name="Pacientes", index=False)
        dictionary_df.to_excel(writer, sheet_name="Diccionario", index=False)

    _format_workbook(output_path, sorted(set(specialty_values or [])))
    return output_path
